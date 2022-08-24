#!/usr/bin/python

import pandas as pd
import sqlite3
import xlsxwriter
from pandas import ExcelWriter
from openpyxl import load_workbook

con = sqlite3.connect('AUDIO.db')
cur = con.cursor()

#Create Excel Workbook
workbook = xlsxwriter.Workbook('tables.xlsx')
worksheet = workbook.add_worksheet()
workbook.close()

#Load excel Workbook using openpyxl
book = load_workbook('tables.xlsx')
writer = ExcelWriter('tables.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

def query_db(query,sheet_name):

    df = pd.read_sql(query,con)
    df.to_excel(writer, sheet_name=sheet_name, index=False)

query_db('select * from `AOD Master` where `Artist` = "Various Artists";',"Various Artists")
query_db('select * from `AOD Master` where `Artist` = "Midnight Oil ";',"Midnight Oil")

# remove empty first_sheet if there is atleast 1 log sheet
if len(book.sheetnames) > 1:
    first_sheet = book['Sheet1']
    book.remove(first_sheet)

writer.save()